from collections import defaultdict
from csv import excel
from io import StringIO, BytesIO
import csv
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, BufferedInputFile
from aiogram.utils.formatting import Text, Code, Italic, Bold
from DataBase.DAO import DAO
from bot_classes import SuperTest, convert_stest_id, generate_custom_qr_code
from handlers.pagination import create_pagination
from aiogram import Router, F, Bot
from handlers.create_test import poosh_msg, get_super_test_in_str
from datetime import datetime, timedelta
from openpyxl import Workbook

router = Router()

# В разделе с импортами добавьте:
from inline_keyboards.tests_menu import get_super_test_actions_keyb, get_cancel_edit_keyb, get_results_var_keyb

def get_score(criteria, num):
    criteria = [int(i) for i in criteria]
    if num <= criteria[1]:
        return 2
    elif num <= criteria[2]:
        return 3
    elif num <= criteria[3]:
        return 3
    else:
        return 5

def stest_to_str(super_test: SuperTest,):
    if super_test.end_date >= datetime.now():
        end_date = ["\nОкончание: ", Code(super_test.end_date.strftime('%d.%m.%Y %H:%M'))]
    else:
        end_date = ["\nТест завершён!"]

    text = Text(f"Супер-тест: ", Code(super_test.name),
        *end_date,
        f"\nВопросов: ", Code(len(super_test.tests_id)),
        f"\nОписание: ",
        Code(super_test.description if super_test.description else 'Отсутствует')).as_markdown()
    return text

async def save_data_end_edit_msg(state: FSMContext, new_state, bot: Bot, data, chat_id: int, text: str,
                                 keyboard: InlineKeyboardMarkup = None):
    """Функция удаляет старое сообщение, отправляет новое, заносит его id в data, и изменяет стейт пользователя"""
    await bot.delete_message(chat_id, data['test_menu_msg_id'])
    bot_msg = await bot.send_message(chat_id, text, reply_markup=keyboard, parse_mode=ParseMode.MARKDOWN_V2)
    data['test_menu_msg_id'] = bot_msg.message_id
    await state.set_state(new_state)
    await state.set_data(data)

@router.message(Command("tests"))
async def show_user_supertests(msg: Message, state: FSMContext):
    """Показать супер-тесты пользователя с пагинацией"""
    user_supertests = DAO.SuperTest.get_for_user(msg.from_user.id)

    if not user_supertests:
        await msg.answer("У вас пока нет созданных супер-тестов")
        return

    texts = [stest.name for stest in user_supertests]
    indexes = [stest.id for stest in user_supertests]

    text, keyb = await create_pagination(
        state=state,
        texts=texts,
        indexes=indexes,
        callback="super_test_",
        back_callback="back_to_menu",
        items_on_page=5
    )

    new_msg = await msg.answer(f"Ваши супер-тесты:\n{text}", reply_markup=keyb)
    data = await state.get_data()
    data['test_menu_msg_id'] = new_msg.message_id
    await state.set_data(data)


# Новый обработчик для выбора супер-теста
@router.callback_query(F.data.startswith("super_test_"))
async def handle_super_test_selection(callback: CallbackQuery, state: FSMContext, bot: Bot):
    test_id = int(callback.data.split("_")[2])
    super_test = DAO.SuperTest.get_for_id(test_id)

    text = stest_to_str(super_test)

    try:
        await callback.message.edit_text(
            text,
            reply_markup=get_super_test_actions_keyb(super_test),
            parse_mode=ParseMode.MARKDOWN_V2
        )
    except:
        data = await state.get_data()
        new_msg = await bot.send_message(callback.message.chat.id, text,
                                          reply_markup=get_super_test_actions_keyb(super_test),
                                          parse_mode=ParseMode.MARKDOWN_V2)
        await callback.message.delete()
        data['test_menu_msg_id'] = new_msg
        await state.set_data(data)


# Обработчик кнопки "Назад"
@router.callback_query(F.data == "back_to_supertests")
async def back_to_supertests(callback: CallbackQuery, state: FSMContext):
    user_supertests = DAO.SuperTest.get_for_user(callback.from_user.id)

    if not user_supertests:
        await callback.message.answer("У вас пока нет созданных супер-тестов")
        return

    texts = [stest.name for stest in user_supertests]
    indexes = [stest.id for stest in user_supertests]

    text, keyb = await create_pagination(
        state=state,
        texts=texts,
        indexes=indexes,
        callback="super_test_",
        back_callback="back_to_menu",
        items_on_page=5
    )

    new_msg = await callback.message.answer(f"Ваши тесты:\n{text}", reply_markup=keyb)
    data = await state.get_data()
    data['test_menu_msg_id'] = new_msg.message_id
    await state.set_data(data)


class Get(StatesGroup):
    EditSuperTestDate = State()
    ExtendSuperTestDate = State()  # Новое состояние


# В роутер добавьте обработчики
@router.callback_query(F.data.startswith("extend_test_"))
async def extend_test_start(calb: CallbackQuery, state: FSMContext, bot: Bot):
    stest_id = int(calb.data.split("_")[2])
    data = await state.get_data()

    await save_data_end_edit_msg(
        state=state,
        new_state=Get.ExtendSuperTestDate,
        bot=bot,
        data=data,
        chat_id=calb.message.chat.id,
        text=Text(
            "Введите дату и время в формате:\n"
            "день:месяц:год:часы:минуты\n"
            "Пример: 04:03:2025:16:42 —>\n"
            "4 марта 2025 года, 16:42"
        ).as_markdown(),
        keyboard=get_cancel_edit_keyb(stest_id)
    )
    await state.update_data(current_stest_id=stest_id)


@router.message(Get.ExtendSuperTestDate)
async def process_extend_date(msg: Message, state: FSMContext, bot: Bot):
    data = await state.get_data()
    stest_id = data['current_stest_id']

    try:
        new_date = datetime.strptime(msg.text, "%d:%m:%Y:%H:%M")
        if new_date <= datetime.now() + timedelta(hours=1):
            await poosh_msg(msg, "Дата должна быть минимум на час больше текущей")
            return

        DAO.SuperTest.edit_end_date(stest_id, new_date)
        super_test = DAO.SuperTest.get_for_id(stest_id)

        await msg.delete()
        await bot.delete_message(msg.chat.id, data['test_menu_msg_id'])

        text = stest_to_str(super_test)
        new_msg = await msg.answer(text, reply_markup=get_super_test_actions_keyb(super_test),
                                   parse_mode=ParseMode.MARKDOWN_V2)

        await state.update_data(test_menu_msg_id=new_msg.message_id)
        await state.set_state(None)

    except ValueError:
        await poosh_msg(msg, "Некорректный формат даты")


@router.callback_query(F.data.startswith("finish_test_"))
async def finish_test_handler(calb: CallbackQuery):
    stest_id = int(calb.data.split("_")[2])

    # Устанавливаем дату окончания на текущее время - 1 час
    new_end_date = datetime.now() - timedelta(hours=1)
    DAO.SuperTest.edit_end_date(stest_id, new_end_date)

    # Обновляем сообщение
    super_test = DAO.SuperTest.get_for_id(stest_id)
    text = stest_to_str(super_test)

    await calb.message.edit_text(
        text,
        reply_markup=get_super_test_actions_keyb(super_test),
        parse_mode=ParseMode.MARKDOWN_V2
    )


# Обработчик кнопки "Ссылка" (заглушка)
@router.callback_query(F.data.startswith("link_"))
async def link_handler(calb: CallbackQuery, bot: Bot, state: FSMContext):
    data = await state.get_data()
    super_test_id = int(calb.data.split("_")[1])
    super_test_hex = convert_stest_id(super_test_id)  # "красивое" представление id теста * на 45348 в 16-ричной СЧ
    # Создаём ссылку на тест и QR-код
    test_link = f"https://t.me/TestsProjectBot?start={super_test_hex}"
    test_link_qr = generate_custom_qr_code(test_link)
    photo = BufferedInputFile(test_link_qr.getvalue(), f"Test-{super_test_hex}.png")
    test_link_qr.close()
    # Возвращаем результат
    await calb.message.delete()
    new_msg = await bot.send_photo(calb.message.chat.id, photo, caption=f"Ссылка: {test_link}", reply_markup=get_cancel_edit_keyb(super_test_id))
    data['test_menu_msg'] = new_msg.message_id
    await state.set_data(data)


# Обработчик кнопки "Результаты"
@router.callback_query(F.data.startswith("results_"))
async def results_handler(calb: CallbackQuery, state: FSMContext):
    stest_id = int(calb.data.split("_")[1])

    await calb.message.edit_text(
        "Выберите формат для просмотра результатов:",
        reply_markup=get_results_var_keyb(stest_id)
    )


@router.callback_query(F.data.startswith("export_excel_"))
async def export_excel_handler(calb: CallbackQuery):
    stest_id = int(calb.data.split("_")[2])
    results = DAO.Answer.get_detailed_answers(stest_id)
    super_test = DAO.SuperTest.get_for_id(stest_id)
    criteria = DAO.User.get_creteria(calb.from_user.id)

    # Получаем все вопросы теста для заголовков
    questions = {}
    for test_id in super_test.tests_id:
        test = DAO.Test.get(test_id)
        if test:
            questions[test_id] = test.text

    wb = Workbook()
    ws = wb.active
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = 'Имя', 'Никнейм', 'результат', 'оценка'

    for i in range(len(results)):
        ws[f'A{i+2}'], ws[f'B{i+2}'] = results[i]['name'], results[i]['username']
        ws[f'C{i+2}'], ws[f'D{i+2}'] = results[i]['total_score'], str(get_score(criteria, results[i]['total_score']))

    excel_bytes = BytesIO()
    wb.save(excel_bytes)  # Записываем файл в поток
    excel_bytes.seek(0)  # Перемещаем курсор в начало

    # Получаем бинарные данные
    binary_data = excel_bytes.getvalue()

    file = BufferedInputFile(
        file=binary_data,
        filename=f"results_{stest_id}.xlsx"
    )

    await calb.message.answer_document(
        document=file,
        caption=f"Результаты теста: {super_test.name}"
    )


@router.callback_query(F.data.startswith("export_msg_"))
async def export_msg_handler(calb: CallbackQuery):
    stest_id = int(calb.data.split("_")[2])
    results = DAO.Answer.get_detailed_answers(stest_id)
    super_test = DAO.SuperTest.get_for_id(stest_id)

    text = Text(
        Bold(f"Результаты теста: {super_test.name}\n"),
        f"Всего участников: {len(results)}\n\n"
    ).as_markdown()

    # Сортируем по убыванию баллов
    results_sorted = sorted(results, key=lambda x: x['total_score'], reverse=True)[:10]

    for i, res in enumerate(results_sorted, 1):
        user_text = Text(f"{i}. {res['name']} (@{res['username']})\n")

        for answer in res['answers']:
            user_text += Text(
                f"Вопрос: {answer['question']}\n"
                f"Ответ: {answer['answers']}\n"
            )

        user_text += Text(f"Всего баллов: {res['total_score']}\n\n")
        text += user_text.as_markdown()

    if len(results) > 10:
        text += Italic("\nПоказаны топ-10 результатов").as_markdown()

    await calb.message.edit_text(
        text,
        parse_mode=ParseMode.MARKDOWN_V2,
        reply_markup=get_results_var_keyb(stest_id)
    )